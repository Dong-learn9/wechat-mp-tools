"""
抖音账号订阅模块
- 管理订阅的用户列表（增删改查、批量导入）
- 下载订阅用户的全部作品（含风控时间间隔）
- 定期扫描增量下载新作品
- 每个用户的作品信息记录到 Excel 文件
- 下载前检查本地文件是否已存在，避免重复下载
"""
import re
import time
import json
import random
import threading
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Blueprint, jsonify, request

from backend.config import DATA_DIR, load_json, save_json
from backend.douyin import (
    DouyinClient, download_media, ensure_douyin_dirs,
    _add_log, _set_task_state, _reset_task_state,
    _task_state, _task_lock, _task_cancel_event,
    DOUYIN_DIR, clean_filename,
)

subscription_bp = Blueprint("douyin_subscription", __name__, url_prefix="/api/douyin/subscription")

# 订阅数据文件
SUBSCRIPTIONS_FILE = DATA_DIR / "douyin_subscriptions.json"

# Excel 数据目录：每个订阅用户一个 Excel 文件，记录作品信息
EXCEL_DATA_DIR = DOUYIN_DIR / "data"

# Excel 表头字段（按用户指定的顺序）
EXCEL_HEADERS = [
    "作品类型", "采集时间", "UID", "SEC_UID", "ID", "作品ID", "作品描述",
    "作品话题", "视频时长", "视频分辨率", "视频高度", "视频宽度", "作品链接",
    "发布时间", "视频URI", "账号昵称", "年龄", "账号签名", "下载地址",
    "音乐作者", "音乐标题", "音乐链接", "静态封面", "动态封面", "隐藏标签",
    "点赞数量", "评论数量", "收藏数量", "分享数量", "播放数量", "额外信息",
]

# 订阅任务状态（独立于普通下载任务，避免冲突）
_sub_task_state = {
    "status": "idle",          # idle / running / paused / completed / failed / cancelled
    "total": 0,
    "current_index": 0,
    "current_title": "",
    "current_user": "",        # 当前正在下载的用户昵称
    "logs": [],
    "downloaded_count": 0,
    "failed_count": 0,
    "skipped_count": 0,        # 增量下载时跳过的已存在作品数
    "mode": "",                # "full" 全量下载 / "incremental" 增量扫描
}
_sub_task_lock = threading.Lock()
_sub_cancel_event = threading.Event()
_sub_pause_event = threading.Event()   # 暂停事件：set=暂停中, clear=运行中


def _wait_if_paused() -> bool:
    """如果处于暂停状态则阻塞等待，直到恢复或取消。
    返回 True 表示等待期间被取消，应中止任务；False 表示可继续执行。"""
    while _sub_pause_event.is_set():
        if _sub_cancel_event.is_set():
            return True
        time.sleep(0.3)
    return False

# 并发下载相关锁
_download_lock = threading.Lock()   # 保护下载计数器
_excel_lock = threading.Lock()      # 保护 Excel 文件写入

# 下载队列（不同作者依次排队执行）
_download_queue = []                # [{"sec_uid": str, "nickname": str, "mode": str}]
_queue_lock = threading.Lock()
_current_sec_uid = ""               # 当前正在下载的 sec_uid

# 默认并发下载数（CDN 文件下载并发，API 调用仍保持串行）
_DEFAULT_CONCURRENCY = 3

# 增量扫描调度器状态
_scan_scheduler_running = False
_scan_scheduler_thread = None
_scan_scheduler_lock = threading.Lock()


def _load_subscriptions() -> list:
    """加载订阅列表"""
    return load_json(SUBSCRIPTIONS_FILE, [])


def _save_subscriptions(subs: list):
    """保存订阅列表"""
    save_json(SUBSCRIPTIONS_FILE, subs)


def _find_subscription(subs: list, sec_uid: str) -> dict:
    """根据 sec_uid 查找订阅"""
    for s in subs:
        if s.get("sec_uid") == sec_uid:
            return s
    return None


def _parse_sec_uid(url: str) -> str:
    """从抖音主页 URL 或分享文本中解析 sec_uid"""
    url = url.strip()
    # 直接是 sec_uid（纯字符串）
    if re.match(r"^[A-Za-z0-9_-]{20,}$", url):
        return url
    # 从 URL 中提取
    match = re.search(r"/user/([A-Za-z0-9_-]+)", url)
    if match:
        return match.group(1)
    return ""


def _resolve_and_parse(raw_url: str) -> dict:
    """解析分享链接，返回 sec_uid 和主页 URL"""
    client = DouyinClient()
    resolved = client.resolve_share_url(raw_url)
    sec_uid = _parse_sec_uid(resolved)
    if not sec_uid:
        # 尝试从原始 URL 解析
        sec_uid = _parse_sec_uid(raw_url)
    return {"sec_uid": sec_uid, "resolved_url": resolved}


def _fetch_user_info(sec_uid: str) -> dict:
    """获取用户详情（昵称、头像等）"""
    client = DouyinClient()
    data = client.get_user_detail(sec_uid)
    user = data.get("user", {})
    return {
        "nickname": user.get("nickname", "未知用户"),
        "avatar": user.get("avatar_thumb", {}).get("url_list", [""])[0] if user.get("avatar_thumb") else "",
        "sec_uid": sec_uid,
        "aweme_count": user.get("aweme_count", 0),
        "follower_count": user.get("follower_count", 0),
    }


def _sub_add_log(message: str):
    """添加订阅任务日志"""
    with _sub_task_lock:
        _sub_task_state["logs"].append(message)
        # 最多保留 300 条
        if len(_sub_task_state["logs"]) > 300:
            _sub_task_state["logs"] = _sub_task_state["logs"][-300:]


def _sub_set_state(**kwargs):
    """更新订阅任务状态"""
    with _sub_task_lock:
        for k, v in kwargs.items():
            _sub_task_state[k] = v


def _sub_reset_state(total: int = 0, mode: str = ""):
    """重置订阅任务状态"""
    _sub_pause_event.clear()
    with _sub_task_lock:
        _sub_task_state.update({
            "status": "running",
            "total": total,
            "current_index": 0,
            "current_title": "",
            "current_user": "",
            "logs": [],
            "downloaded_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "mode": mode,
        })


# ── Excel 记录与文件存在检查 ──────────────────────────────

def _extract_aweme_info(detail: dict, media_info: dict, local_path: str) -> list:
    """从 aweme detail 中提取所有字段，返回与 EXCEL_HEADERS 对应的行数据"""
    author = detail.get("author") or {}
    video = detail.get("video") or {}
    music = detail.get("music") or {}
    stats = detail.get("statistics") or {}
    text_extra = detail.get("text_extra") or []

    aweme_id = detail.get("aweme_id", "")
    # 作品话题：text_extra 中 type=1 的是话题标签
    hashtags = [te.get("hashtag_name", "") for te in text_extra if te.get("type") == 1 and te.get("hashtag_name")]
    # 隐藏标签：text_extra 中 type=0 的是提及/隐藏标签
    hidden_tags = [te.get("hashtag_name", te.get("sec_uid", "")) for te in text_extra if te.get("type") == 0]

    # 发布时间
    create_time = detail.get("create_time", 0)
    try:
        publish_time = datetime.fromtimestamp(int(create_time)).strftime("%Y-%m-%d %H:%M:%S") if create_time else ""
    except Exception:
        publish_time = ""

    # 视频时长（毫秒 → 秒）
    duration_ms = video.get("duration", 0)
    duration_sec = round(int(duration_ms) / 1000, 1) if duration_ms else 0

    # 视频分辨率
    v_width = video.get("width", 0)
    v_height = video.get("height", 0)
    resolution = f"{v_width}x{v_height}" if v_width and v_height else ""

    # 视频 URI
    play_addr = video.get("play_addr") or {}
    video_uri = play_addr.get("uri", "")

    # 封面
    cover = video.get("cover") or {}
    static_cover = (cover.get("url_list") or [""])[0] if cover else ""
    dynamic_cover_obj = video.get("dynamic_cover") or {}
    dynamic_cover = (dynamic_cover_obj.get("url_list") or [""])[0] if dynamic_cover_obj else ""

    # 音乐链接
    play_url = music.get("play_url") or {}
    music_url = (play_url.get("url_list") or [""])[0] if play_url else music.get("uri", "")

    # 作品类型
    item_type = "图文" if media_info.get("type") == "image" else "视频"

    # 额外信息（JSON 字符串）
    extra_info = json.dumps({
        "aweme_type": detail.get("aweme_type", ""),
        "is_top": detail.get("is_top", 0),
        "region": detail.get("region", ""),
        "duration": detail.get("duration", 0),
    }, ensure_ascii=False)

    return [
        item_type,                                    # 作品类型
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"), # 采集时间
        author.get("uid", ""),                        # UID
        author.get("sec_uid", ""),                    # SEC_UID
        "",                                           # ID（行号，由 _append_to_excel 填充）
        aweme_id,                                     # 作品ID
        detail.get("desc", ""),                       # 作品描述
        " ".join(hashtags),                           # 作品话题
        duration_sec,                                 # 视频时长
        resolution,                                   # 视频分辨率
        v_height,                                     # 视频高度
        v_width,                                      # 视频宽度
        f"https://www.douyin.com/video/{aweme_id}",   # 作品链接
        publish_time,                                 # 发布时间
        video_uri,                                    # 视频URI
        author.get("nickname", ""),                   # 账号昵称
        author.get("age", ""),                        # 年龄
        author.get("signature", ""),                  # 账号签名
        local_path,                                   # 下载地址
        music.get("author", ""),                      # 音乐作者
        music.get("title", ""),                       # 音乐标题
        music_url,                                    # 音乐链接
        static_cover,                                 # 静态封面
        dynamic_cover,                                # 动态封面
        " ".join(hidden_tags),                        # 隐藏标签
        stats.get("digg_count", 0),                   # 点赞数量
        stats.get("comment_count", 0),                # 评论数量
        stats.get("collect_count", 0),                # 收藏数量
        stats.get("share_count", 0),                  # 分享数量
        stats.get("play_count", 0),                   # 播放数量
        extra_info,                                   # 额外信息
    ]


def _append_to_excel(nickname: str, row_data: list):
    """将一行作品信息追加到对应用户的 Excel 文件中"""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        _sub_add_log("⚠️ openpyxl 未安装，无法写入 Excel")
        return

    EXCEL_DATA_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = clean_filename(nickname) if nickname else "未知用户"
    excel_path = EXCEL_DATA_DIR / f"{safe_name}.xlsx"

    try:
        if excel_path.exists():
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "作品信息"
            ws.append(EXCEL_HEADERS)

        # 检查是否已存在相同作品ID（避免重复记录）
        aweme_id = row_data[5]  # 作品ID 列索引
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
            if row[0] == aweme_id:
                # 已存在，跳过
                return

        # 填充 ID 列（行号）
        row_data[4] = ws.max_row  # ID 列索引为 4
        ws.append(row_data)
        wb.save(excel_path)
    except Exception as e:
        _sub_add_log(f"⚠️ 写入 Excel 失败: {e}")


def _check_file_exists(media_info: dict, target_dir: Path) -> bool:
    """检查作品文件是否已存在于本地"""
    aweme_id = media_info.get("aweme_id", "")
    title = media_info.get("title", "")
    nickname = media_info.get("nickname", "未知用户")
    item_type = media_info.get("type", "video")
    create_time = media_info.get("create_time", 0)

    if not aweme_id:
        return False

    # 文件命名规则：作者-发布日期-作品类型(视频或者图集)-标题
    try:
        publish_date = datetime.fromtimestamp(int(create_time)).strftime("%Y-%m-%d") if create_time else "未知日期"
    except (ValueError, OSError, OverflowError):
        publish_date = "未知日期"
    type_label = "视频" if item_type == "video" else "图集"
    file_name = clean_filename(f"{nickname}-{publish_date}-{type_label}-{title}")

    user_dir = target_dir / nickname

    if item_type == "video":
        save_file = user_dir / f"{file_name}.mp4"
        return save_file.exists() and save_file.stat().st_size > 0
    else:
        # 图文：检查目录是否存在且有内容
        folder_path = user_dir / file_name
        if not folder_path.exists():
            return False
        # 检查目录下是否有图片文件
        return any(folder_path.glob("*.jpeg")) or any(folder_path.glob("*.jpg"))


# ── 订阅管理 API ──────────────────────────────────────────

@subscription_bp.route("/list", methods=["GET"])
def list_subscriptions():
    """获取所有订阅列表"""
    subs = _load_subscriptions()
    return jsonify({"subscriptions": subs, "count": len(subs)})


@subscription_bp.route("/add", methods=["POST"])
def add_subscription():
    """添加单个订阅（通过主页链接或 sec_uid）"""
    data = request.get_json() or {}
    raw_url = data.get("url", "").strip()

    if not raw_url:
        return jsonify({"error": "请提供主页链接或 sec_uid"}), 400

    # 解析 sec_uid
    try:
        result = _resolve_and_parse(raw_url)
    except Exception as e:
        return jsonify({"error": f"链接解析失败: {str(e)}"}), 400

    sec_uid = result["sec_uid"]
    if not sec_uid:
        return jsonify({"error": "未能从链接解析出 sec_uid，请确认主页格式正确"}), 400

    subs = _load_subscriptions()

    # 检查是否已订阅
    existing = _find_subscription(subs, sec_uid)
    if existing:
        return jsonify({"error": "该用户已在订阅列表中", "subscription": existing}), 400

    # 获取用户详情
    try:
        info = _fetch_user_info(sec_uid)
    except Exception as e:
        # 获取详情失败时仍可添加，使用默认信息
        info = {"nickname": "未知用户", "avatar": "", "sec_uid": sec_uid, "aweme_count": 0, "follower_count": 0}

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sub = {
        "sec_uid": sec_uid,
        "nickname": info["nickname"],
        "avatar": info["avatar"],
        "homepage_url": f"https://www.douyin.com/user/{sec_uid}",
        "aweme_count": info["aweme_count"],
        "follower_count": info["follower_count"],
        "subscribed_at": now,
        "last_download_at": "",           # 上次下载时间
        "last_download_count": 0,         # 上次下载的作品数
        "downloaded_aweme_ids": [],       # 已下载的作品 ID 列表（用于增量判断）
        "auto_scan": True,                # 是否参与增量扫描
        "status": "idle",                 # idle / downloading / completed / failed
    }

    subs.append(sub)
    _save_subscriptions(subs)

    return jsonify({"message": f"已订阅: {info['nickname']}", "subscription": sub})


@subscription_bp.route("/add-batch", methods=["POST"])
def add_batch_subscriptions():
    """批量导入订阅（通过文本，每行一个主页链接）"""
    data = request.get_json() or {}
    text = data.get("text", "")

    if not text.strip():
        return jsonify({"error": "请提供至少一行主页链接"}), 400

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return jsonify({"error": "未解析到有效链接"}), 400

    subs = _load_subscriptions()
    results = {"success": [], "failed": [], "duplicated": []}

    for line in lines:
        try:
            result = _resolve_and_parse(line)
            sec_uid = result["sec_uid"]
            if not sec_uid:
                results["failed"].append({"url": line, "error": "无法解析 sec_uid"})
                continue

            if _find_subscription(subs, sec_uid):
                results["duplicated"].append({"url": line, "sec_uid": sec_uid})
                continue

            try:
                info = _fetch_user_info(sec_uid)
            except Exception:
                info = {"nickname": "未知用户", "avatar": "", "sec_uid": sec_uid, "aweme_count": 0, "follower_count": 0}

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sub = {
                "sec_uid": sec_uid,
                "nickname": info["nickname"],
                "avatar": info["avatar"],
                "homepage_url": f"https://www.douyin.com/user/{sec_uid}",
                "aweme_count": info["aweme_count"],
                "follower_count": info["follower_count"],
                "subscribed_at": now,
                "last_download_at": "",
                "last_download_count": 0,
                "downloaded_aweme_ids": [],
                "auto_scan": True,
                "status": "idle",
            }
            subs.append(sub)
            results["success"].append({"url": line, "nickname": info["nickname"]})

            # 批量导入时每个用户之间加延迟，避免风控
            time.sleep(random.uniform(1.0, 2.0))

        except Exception as e:
            results["failed"].append({"url": line, "error": str(e)})

    _save_subscriptions(subs)
    return jsonify({
        "message": f"导入完成: 成功 {len(results['success'])}，重复 {len(results['duplicated'])}，失败 {len(results['failed'])}",
        "results": results,
    })


@subscription_bp.route("/remove", methods=["POST"])
def remove_subscription():
    """移除订阅"""
    data = request.get_json() or {}
    sec_uid = data.get("sec_uid", "").strip()

    if not sec_uid:
        return jsonify({"error": "请提供 sec_uid"}), 400

    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if not target:
        return jsonify({"error": "未找到该订阅"}), 404

    subs = [s for s in subs if s.get("sec_uid") != sec_uid]
    _save_subscriptions(subs)

    return jsonify({"message": f"已移除订阅: {target.get('nickname', '')}"})


@subscription_bp.route("/clear-downloaded", methods=["POST"])
def clear_downloaded_records():
    """清空指定订阅的已下载记录，以便重新下载已删除的作品"""
    data = request.get_json() or {}
    sec_uid = data.get("sec_uid", "").strip()

    if not sec_uid:
        return jsonify({"error": "请提供 sec_uid"}), 400

    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if not target:
        return jsonify({"error": "未找到该订阅"}), 404

    cleared_count = len(target.get("downloaded_aweme_ids", []))
    target["downloaded_aweme_ids"] = []
    target["last_download_count"] = 0
    _save_subscriptions(subs)

    nickname = target.get("nickname", "")
    _sub_add_log(f"🗑️ 已清空 {nickname} 的已下载记录（{cleared_count} 项）")
    return jsonify({
        "message": f"已清空 {nickname} 的已下载记录（{cleared_count} 项），可重新下载",
        "cleared_count": cleared_count
    })


@subscription_bp.route("/toggle-scan", methods=["POST"])
def toggle_auto_scan():
    """开启/关闭某个订阅的自动增量扫描"""
    data = request.get_json() or {}
    sec_uid = data.get("sec_uid", "").strip()
    enabled = bool(data.get("enabled", True))

    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if not target:
        return jsonify({"error": "未找到该订阅"}), 404

    target["auto_scan"] = enabled
    _save_subscriptions(subs)

    return jsonify({"message": f"已{'开启' if enabled else '关闭'}自动扫描: {target.get('nickname', '')}"})


@subscription_bp.route("/refresh", methods=["POST"])
def refresh_subscription_info():
    """刷新订阅用户的最新信息（昵称、作品数等）"""
    subs = _load_subscriptions()
    updated = 0

    for sub in subs:
        try:
            info = _fetch_user_info(sub["sec_uid"])
            sub["nickname"] = info["nickname"]
            sub["avatar"] = info["avatar"]
            sub["aweme_count"] = info["aweme_count"]
            sub["follower_count"] = info["follower_count"]
            updated += 1
            time.sleep(random.uniform(1.0, 2.0))
        except Exception:
            continue

    _save_subscriptions(subs)
    return jsonify({"message": f"已刷新 {updated}/{len(subs)} 个订阅的信息"})


# ── 订阅下载 API ──────────────────────────────────────────

@subscription_bp.route("/download", methods=["POST"])
def download_subscription():
    """下载指定订阅用户的全部作品（全量）。
    如果已有任务在运行，则加入下载队列排队等待。"""
    data = request.get_json() or {}
    sec_uid = data.get("sec_uid", "").strip()

    if not sec_uid:
        return jsonify({"error": "请提供 sec_uid"}), 400

    # 检查同一作者是否正在下载或已在队列中
    if _is_downloading_or_queued(sec_uid):
        return jsonify({"error": "该作者正在下载或已在队列中，请勿重复添加"}), 400

    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if not target:
        return jsonify({"error": "未找到该订阅"}), 404

    nickname = target.get("nickname", "")

    # 如果当前没有运行中的任务，立即启动
    if _sub_task_state["status"] not in ("running", "paused"):
        _sub_cancel_event.clear()
        _sub_pause_event.clear()
        thread = threading.Thread(
            target=_run_queued_download,
            args=(sec_uid, nickname, "full"),
            daemon=True,
        )
        thread.start()
        return jsonify({"message": f"已开始下载 {nickname} 的全部作品"})

    # 否则加入队列排队
    with _queue_lock:
        _download_queue.append({"sec_uid": sec_uid, "nickname": nickname, "mode": "full"})
        position = len(_download_queue)

    _sub_add_log(f"📋 {nickname} 已加入下载队列（第 {position} 位）")
    return jsonify({
        "message": f"{nickname} 已加入下载队列（排第 {position} 位）",
        "queued": True,
        "queue_position": position
    })


@subscription_bp.route("/scan-all", methods=["POST"])
def scan_all_subscriptions():
    """增量扫描所有开启自动扫描的订阅用户，下载新作品"""
    if _sub_task_state["status"] in ("running", "paused"):
        return jsonify({"error": "订阅下载任务正在运行或暂停中，请先取消或等待完成"}), 400

    subs = _load_subscriptions()
    targets = [s for s in subs if s.get("auto_scan", True)]

    if not targets:
        return jsonify({"error": "没有开启自动扫描的订阅"}), 400

    # 启动后台增量扫描
    _sub_cancel_event.clear()
    _sub_pause_event.clear()
    thread = threading.Thread(
        target=_run_queued_scan,
        args=(targets,),
        daemon=True,
    )
    thread.start()

    return jsonify({"message": f"已开始增量扫描 {len(targets)} 个订阅用户"})


@subscription_bp.route("/cancel", methods=["POST"])
def cancel_subscription_download():
    """取消订阅下载任务"""
    _sub_cancel_event.set()
    _sub_pause_event.clear()  # 解除暂停以便等待循环退出
    _sub_set_state(status="cancelled")
    return jsonify({"message": "已发送取消信号"})


@subscription_bp.route("/queue", methods=["GET"])
def get_download_queue():
    """获取下载队列"""
    with _queue_lock:
        return jsonify({
            "queue": list(_download_queue),
            "count": len(_download_queue),
            "current_sec_uid": _current_sec_uid,
        })


@subscription_bp.route("/queue/remove", methods=["POST"])
def remove_from_queue():
    """从下载队列中移除指定作者"""
    data = request.get_json() or {}
    sec_uid = data.get("sec_uid", "").strip()
    if not sec_uid:
        return jsonify({"error": "请提供 sec_uid"}), 400
    with _queue_lock:
        before = len(_download_queue)
        _download_queue[:] = [item for item in _download_queue if item["sec_uid"] != sec_uid]
        after = len(_download_queue)
    removed = before - after
    if removed > 0:
        _sub_add_log(f"📋 已从队列中移除 {removed} 项")
        return jsonify({"message": "已从队列中移除"})
    return jsonify({"message": "队列中未找到该项"})


@subscription_bp.route("/queue/clear", methods=["POST"])
def clear_download_queue():
    """清空下载队列"""
    with _queue_lock:
        count = len(_download_queue)
        _download_queue.clear()
    if count > 0:
        _sub_add_log(f"📋 已清空下载队列（{count} 项）")
    return jsonify({"message": f"已清空下载队列（{count} 项）"})


@subscription_bp.route("/pause", methods=["POST"])
def pause_subscription_download():
    """暂停订阅下载任务"""
    if _sub_task_state["status"] != "running":
        return jsonify({"error": "当前没有正在运行的下载任务"}), 400
    _sub_pause_event.set()
    _sub_set_state(status="paused")
    _sub_add_log("⏸️ 下载已暂停")
    return jsonify({"message": "下载已暂停"})


@subscription_bp.route("/resume", methods=["POST"])
def resume_subscription_download():
    """恢复订阅下载任务"""
    if _sub_task_state["status"] != "paused":
        return jsonify({"error": "当前任务未处于暂停状态"}), 400
    _sub_pause_event.clear()
    _sub_set_state(status="running")
    _sub_add_log("▶️ 下载已恢复")
    return jsonify({"message": "下载已恢复"})


@subscription_bp.route("/progress", methods=["GET"])
def get_subscription_progress():
    """获取订阅下载进度（含队列信息）"""
    with _sub_task_lock:
        state = dict(_sub_task_state)
    state["concurrency"] = _get_download_concurrency()
    with _queue_lock:
        state["queue"] = list(_download_queue)
        state["queue_count"] = len(_download_queue)
        state["current_sec_uid"] = _current_sec_uid
    return jsonify(state)


@subscription_bp.route("/concurrency", methods=["GET", "POST"])
def download_concurrency():
    """获取或设置下载并发数（1-10）"""
    if request.method == "GET":
        return jsonify({"concurrency": _get_download_concurrency()})

    data = request.get_json() or {}
    try:
        n = int(data.get("concurrency", _DEFAULT_CONCURRENCY))
    except (ValueError, TypeError):
        n = _DEFAULT_CONCURRENCY
    n = max(1, min(10, n))
    _set_download_concurrency(n)
    return jsonify({"concurrency": n, "message": f"下载并发数已设置为 {n}"})


@subscription_bp.route("/scheduler-status", methods=["GET"])
def get_scheduler_status():
    """获取增量扫描调度器状态"""
    return jsonify({
        "running": _scan_scheduler_running,
        "interval_minutes": _get_scan_interval(),
    })


@subscription_bp.route("/scheduler-toggle", methods=["POST"])
def toggle_scheduler():
    """开启/关闭增量扫描调度器"""
    global _scan_scheduler_running, _scan_scheduler_thread

    data = request.get_json() or {}
    enabled = bool(data.get("enabled", False))
    interval = int(data.get("interval_minutes", 60))

    _set_scan_interval(interval)

    if enabled:
        if not _scan_scheduler_running:
            _scan_scheduler_running = True
            _scan_scheduler_thread = threading.Thread(
                target=_run_scan_scheduler, daemon=True
            )
            _scan_scheduler_thread.start()
            return jsonify({"message": f"增量扫描调度器已启动，间隔 {interval} 分钟"})
        return jsonify({"message": "调度器已在运行中"})
    else:
        _scan_scheduler_running = False
        return jsonify({"message": "增量扫描调度器已停止"})


# 调度器间隔配置存储在 app_settings.json
def _get_scan_interval() -> int:
    """获取扫描间隔（分钟）"""
    from backend.config import get_settings, save_settings
    settings = get_settings()
    return int(settings.get("douyin_scan_interval_minutes", 60))


def _set_scan_interval(minutes: int):
    """设置扫描间隔（分钟）"""
    from backend.config import get_settings, save_settings
    settings = get_settings()
    settings["douyin_scan_interval_minutes"] = max(5, minutes)
    save_settings(settings)


def _get_download_concurrency() -> int:
    """获取下载并发数"""
    from backend.config import get_settings
    settings = get_settings()
    return max(1, min(10, int(settings.get("douyin_download_concurrency", _DEFAULT_CONCURRENCY))))


def _set_download_concurrency(n: int):
    """设置下载并发数（1-10）"""
    from backend.config import get_settings, save_settings
    settings = get_settings()
    settings["douyin_download_concurrency"] = max(1, min(10, n))
    save_settings(settings)


# ── 下载队列管理 ──────────────────────────────────────────

def _is_downloading_or_queued(sec_uid: str) -> bool:
    """检查某 sec_uid 是否正在下载或已在队列中"""
    with _queue_lock:
        if _current_sec_uid == sec_uid and _sub_task_state["status"] in ("running", "paused"):
            return True
        for item in _download_queue:
            if item["sec_uid"] == sec_uid:
                return True
    return False


def _start_next_queued():
    """启动队列中的下一个下载任务"""
    global _current_sec_uid
    with _queue_lock:
        if not _download_queue:
            _current_sec_uid = ""
            return
        next_item = _download_queue.pop(0)
        _current_sec_uid = next_item["sec_uid"]

    sec_uid = next_item["sec_uid"]
    nickname = next_item["nickname"]
    mode = next_item.get("mode", "full")

    _sub_cancel_event.clear()
    _sub_pause_event.clear()
    _sub_add_log(f"▶️ 开始下载队列中的下一个: {nickname}")

    thread = threading.Thread(
        target=_run_queued_download,
        args=(sec_uid, nickname, mode),
        daemon=True,
    )
    thread.start()


def _run_queued_download(sec_uid: str, nickname: str, mode: str):
    """下载包装器：执行下载后自动启动队列中的下一个任务"""
    global _current_sec_uid
    _current_sec_uid = sec_uid
    try:
        _run_subscription_download(sec_uid, nickname, mode)
    finally:
        _current_sec_uid = ""
        _start_next_queued()


def _run_queued_scan(targets: list):
    """增量扫描包装器：执行扫描后自动启动队列中的下一个任务"""
    try:
        _run_incremental_scan(targets)
    finally:
        _start_next_queued()


def _run_scan_scheduler():
    """增量扫描调度器主循环"""
    global _scan_scheduler_running

    while _scan_scheduler_running:
        try:
            # 检查是否有任务正在运行或暂停中
            if _sub_task_state["status"] not in ("running", "paused"):
                _sub_add_log("调度器: 开始执行增量扫描...")
                subs = _load_subscriptions()
                targets = [s for s in subs if s.get("auto_scan", True)]
                if targets:
                    _run_incremental_scan(targets)
        except Exception as e:
            _sub_add_log(f"调度器异常: {e}")

        # 等待下一次扫描
        interval = _get_scan_interval() * 60
        for _ in range(interval):
            if not _scan_scheduler_running:
                return
            time.sleep(1)


# ── 后台下载任务执行器 ────────────────────────────────────

def _download_one_work(idx, item, media_info, local_path, nickname, total):
    """在工作线程中下载单个作品，返回结果字典或抛出异常。含重试机制。"""
    if _sub_cancel_event.is_set():
        return None

    # 暂停检查
    if _wait_if_paused():
        return None

    title_preview = media_info["title"][:30]
    max_retries = 3  # 最大重试次数
    last_error = None

    for attempt in range(1, max_retries + 1):
        if _sub_cancel_event.is_set():
            return None
        # 每次重试前也检查暂停
        if _wait_if_paused():
            return None
        try:
            if attempt == 1:
                _sub_add_log(f"[{idx}/{total}] 下载: {title_preview}...")
            else:
                _sub_add_log(f"[{idx}/{total}] 第 {attempt} 次重试: {title_preview}...")

            result = download_media(media_info, DOUYIN_DIR)

            # 写入 Excel（线程安全）
            row_data = _extract_aweme_info(item, media_info, local_path)
            with _excel_lock:
                _append_to_excel(nickname, row_data)

            return {
                "aweme_id": item.get("aweme_id", ""),
                "title": result["title"],
            }
        except Exception as e:
            last_error = e
            if attempt < max_retries:
                retry_delay = random.uniform(2.0, 4.0)
                _sub_add_log(f"[{idx}/{total}] ⚠️ 下载失败（第 {attempt} 次），{retry_delay:.1f} 秒后重试: {e}")
                time.sleep(retry_delay)
            else:
                _sub_add_log(f"[{idx}/{total}] ❌ 下载失败（已重试 {max_retries} 次）: {e}")

    raise last_error


def _run_subscription_download(sec_uid: str, nickname: str, mode: str = "full"):
    """下载单个订阅用户的全部作品"""
    _sub_reset_state(total=0, mode=mode)
    _sub_add_log(f"开始下载订阅用户: {nickname} (sec_uid: {sec_uid[:20]}...)")

    client = DouyinClient()
    ensure_douyin_dirs()

    try:
        # 加载已下载的作品 ID（用于增量判断）
        subs = _load_subscriptions()
        sub = _find_subscription(subs, sec_uid)
        downloaded_ids = set(sub.get("downloaded_aweme_ids", [])) if sub else set()

        # 第一步：收集作品列表
        _sub_add_log("正在获取用户作品列表...")
        all_items = []
        cursor = 0
        page = 0
        max_pages = 50  # 全量下载最多 50 页

        while page < max_pages:
            if _sub_cancel_event.is_set():
                _sub_add_log("⚠️ 用户取消了任务")
                _sub_set_state(status="cancelled")
                _update_sub_status(sec_uid, "idle")
                return

            # 暂停检查（页面抓取阶段也响应暂停）
            if _wait_if_paused():
                _sub_add_log("⚠️ 暂停期间被取消")
                _sub_set_state(status="cancelled")
                _update_sub_status(sec_uid, "idle")
                return

            page += 1
            _sub_add_log(f"正在获取第 {page} 页作品 (cursor={cursor})...")

            try:
                aweme_list, next_cursor, has_more = client.get_user_videos(sec_uid, cursor)
            except Exception as e:
                _sub_add_log(f"⚠️ 获取第 {page} 页失败: {e}")
                break

            if not aweme_list:
                _sub_add_log(f"第 {page} 页返回空数据，已到达最后一页")
                break

            all_items.extend(aweme_list)
            _sub_add_log(f"第 {page} 页获取 {len(aweme_list)} 个，累计 {len(all_items)} 个")

            if not has_more:
                _sub_add_log("已获取全部作品")
                break

            cursor = next_cursor
            # 风控延迟
            delay = random.uniform(1.5, 4.0)
            _sub_add_log(f"休眠 {delay:.1f} 秒规避风控...")
            time.sleep(delay)

        if not all_items:
            _sub_set_state(status="failed")
            _sub_add_log("❌ 未获取到任何作品")
            _update_sub_status(sec_uid, "failed")
            return

        # 过滤已下载的作品（增量模式）
        if mode == "incremental" and downloaded_ids:
            new_items = [it for it in all_items if it.get("aweme_id") not in downloaded_ids]
            skipped = len(all_items) - len(new_items)
            _sub_add_log(f"增量扫描: 共 {len(all_items)} 个作品，已下载 {skipped} 个，新增 {len(new_items)} 个")
            all_items = new_items
            _sub_set_state(skipped_count=skipped)
        else:
            _sub_add_log(f"🚀 共 {len(all_items)} 个作品待下载")

        if not all_items:
            _sub_set_state(status="completed")
            _sub_add_log("✅ 没有新作品需要下载")
            _update_sub_status(sec_uid, "completed")
            _update_sub_download_info(sec_uid, len(downloaded_ids))
            return

        _sub_set_state(total=len(all_items))
        _sub_set_state(current_user=nickname)

        # 第二步：串行解析作品信息 + 检查本地是否已存在（无网络请求，速度快）
        downloaded = 0
        failed = 0
        skipped_local = 0  # 本地已存在而跳过的数量
        new_downloaded_ids = []
        download_tasks = []  # 待下载的任务列表: (idx, item, media_info, local_path)

        _sub_add_log("正在解析作品信息并检查本地文件...")
        for idx, item in enumerate(all_items, 1):
            if _sub_cancel_event.is_set():
                _sub_add_log("⚠️ 用户取消了任务")
                _sub_set_state(status="cancelled")
                _update_sub_status(sec_uid, "idle")
                return

            # 暂停检查
            if _wait_if_paused():
                _sub_add_log("⚠️ 暂停期间被取消")
                _sub_set_state(status="cancelled")
                _update_sub_status(sec_uid, "idle")
                return

            try:
                media_info = DouyinClient.parse_media_info(item)
                if not media_info["urls"]:
                    _sub_add_log(f"⚠️ 第 {idx} 项无可用资源，跳过")
                    failed += 1
                    _sub_set_state(failed_count=failed)
                    continue

                # 计算本地保存路径（与 download_media 的命名逻辑一致）
                # 文件命名规则：作者-发布日期-作品类型(视频或者图集)-标题
                ct = media_info.get("create_time", 0)
                try:
                    pd = datetime.fromtimestamp(int(ct)).strftime("%Y-%m-%d") if ct else "未知日期"
                except (ValueError, OSError, OverflowError):
                    pd = "未知日期"
                tl = "视频" if media_info["type"] == "video" else "图集"
                fn = clean_filename(f"{media_info.get('nickname', '未知用户')}-{pd}-{tl}-{media_info['title']}")
                user_dir = DOUYIN_DIR / media_info.get("nickname", "未知用户")
                if media_info["type"] == "video":
                    local_path = str(user_dir / f"{fn}.mp4")
                else:
                    local_path = str(user_dir / fn)

                # 检查本地文件是否已存在，存在则跳过下载
                if _check_file_exists(media_info, DOUYIN_DIR):
                    skipped_local += 1
                    _sub_set_state(skipped_count=_sub_task_state.get("skipped_count", 0) + 1)
                    _sub_add_log(f"[{idx}/{len(all_items)}] ⏭ 本地已存在，跳过: {media_info['title'][:30]}...")
                    # 仍然记录到 Excel（确保信息完整）
                    row_data = _extract_aweme_info(item, media_info, local_path)
                    with _excel_lock:
                        _append_to_excel(nickname, row_data)
                    new_downloaded_ids.append(item.get("aweme_id", ""))
                    continue

                download_tasks.append((idx, item, media_info, local_path))
            except Exception as e:
                failed += 1
                _sub_set_state(failed_count=failed)
                _sub_add_log(f"❌ 第 {idx} 项解析失败: {e}")

        _sub_add_log(f"解析完成: 需下载 {len(download_tasks)} 个，本地已存在 {skipped_local} 个，解析失败 {failed} 个")

        # 第三步：多线程并发下载（CDN 文件下载，风控较松）
        if download_tasks and not _sub_cancel_event.is_set():
            concurrency = _get_download_concurrency()
            _sub_add_log(f"🚀 开始并发下载（{concurrency} 线程）...")
            total_count = len(all_items)

            # 按批次并发下载，每批 concurrency 个，批次间加短延迟
            batch_size = concurrency
            for batch_start in range(0, len(download_tasks), batch_size):
                if _sub_cancel_event.is_set():
                    _sub_add_log("⚠️ 用户取消了下载")
                    break

                # 暂停检查（批次间）
                if _wait_if_paused():
                    _sub_add_log("⚠️ 暂停期间被取消")
                    break

                batch = download_tasks[batch_start:batch_start + batch_size]

                with ThreadPoolExecutor(max_workers=len(batch)) as executor:
                    futures = {}
                    for idx, item, media_info, local_path in batch:
                        if _sub_cancel_event.is_set():
                            break
                        future = executor.submit(
                            _download_one_work, idx, item, media_info,
                            local_path, nickname, total_count
                        )
                        futures[future] = idx

                    for future in as_completed(futures):
                        if _sub_cancel_event.is_set():
                            break
                        try:
                            result = future.result()
                            if result:
                                with _download_lock:
                                    downloaded += 1
                                    new_downloaded_ids.append(result["aweme_id"])
                                    _sub_set_state(
                                        downloaded_count=downloaded,
                                        current_title=result["title"],
                                        current_index=futures[future],
                                    )
                        except Exception as e:
                            with _download_lock:
                                failed += 1
                                _sub_set_state(failed_count=failed)
                            _sub_add_log(f"❌ 下载失败: {e}")

                # 批次间风控延迟（并发下载后短延迟，比串行逐个延迟短得多）
                if batch_start + batch_size < len(download_tasks):
                    if _sub_cancel_event.is_set():
                        break
                    delay = random.uniform(1.0, 2.5)
                    _sub_add_log(f"批次间隔休眠 {delay:.1f} 秒...")
                    time.sleep(delay)

        _sub_add_log(f"🎉 下载完成! 成功 {downloaded}，失败 {failed}，本地已存在跳过 {skipped_local}")
        _sub_set_state(status="completed")

        # 更新订阅状态
        _update_sub_status(sec_uid, "completed")
        _update_sub_download_info(sec_uid, len(downloaded_ids) + downloaded + skipped_local)

        # 更新已下载作品 ID 列表
        all_ids = list(downloaded_ids) + new_downloaded_ids
        _update_sub_downloaded_ids(sec_uid, all_ids)

    except Exception as e:
        _sub_add_log(f"💥 下载任务异常: {e}")
        _sub_set_state(status="failed")
        _update_sub_status(sec_uid, "failed")


def _run_incremental_scan(targets: list):
    """增量扫描多个订阅用户"""
    _sub_reset_state(total=len(targets), mode="incremental")
    _sub_add_log(f"开始增量扫描 {len(targets)} 个订阅用户...")

    for idx, sub in enumerate(targets, 1):
        if _sub_cancel_event.is_set():
            _sub_add_log("⚠️ 用户取消了扫描")
            _sub_set_state(status="cancelled")
            return

        # 暂停检查
        if _wait_if_paused():
            _sub_add_log("⚠️ 暂停期间被取消")
            _sub_set_state(status="cancelled")
            return

        sec_uid = sub.get("sec_uid", "")
        nickname = sub.get("nickname", "未知用户")
        _sub_set_state(current_index=idx, current_user=nickname)
        _sub_add_log(f"[{idx}/{len(targets)}] 扫描用户: {nickname}")

        try:
            _run_subscription_download(sec_uid, nickname, "incremental")
        except Exception as e:
            _sub_add_log(f"❌ 扫描 {nickname} 失败: {e}")

        # 用户之间的间隔
        if idx < len(targets):
            delay = random.uniform(5.0, 10.0)
            _sub_add_log(f"用户间隔休眠 {delay:.1f} 秒...")
            time.sleep(delay)

    _sub_add_log("🎉 增量扫描全部完成!")
    _sub_set_state(status="completed")


def _update_sub_status(sec_uid: str, status: str):
    """更新订阅状态"""
    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if target:
        target["status"] = status
        _save_subscriptions(subs)


def _update_sub_download_info(sec_uid: str, total_downloaded: int):
    """更新订阅的下载信息"""
    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if target:
        target["last_download_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        target["last_download_count"] = total_downloaded
        _save_subscriptions(subs)


def _update_sub_downloaded_ids(sec_uid: str, ids: list):
    """更新已下载作品 ID 列表"""
    subs = _load_subscriptions()
    target = _find_subscription(subs, sec_uid)
    if target:
        # 最多保留 2000 个 ID，避免文件过大
        target["downloaded_aweme_ids"] = ids[-2000:] if len(ids) > 2000 else ids
        _save_subscriptions(subs)
